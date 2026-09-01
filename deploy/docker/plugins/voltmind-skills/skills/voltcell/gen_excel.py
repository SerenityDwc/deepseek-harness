#!/usr/bin/env python3
"""VoltCell v5.1: 由设计计算 JSON 生成与 BDA1.0-2.0 同构的 Excel 设计表。
用法: python gen_excel.py '<design-json>' '<output.xlsx>'
设计 JSON = script.py 输出（含 params41.sections 完整参数集）。"""
import json
import sys
from openpyxl import Workbook


def main(argv):
    if len(argv) < 2:
        print(json.dumps({"error": "need <design-json> <output.xlsx>"}), file=sys.stderr)
        return 2
    try:
        design = json.loads(argv[0])
        out = argv[1]
    except json.JSONDecodeError as exc:
        print(json.dumps({"error": f"bad design json: {exc}"}), file=sys.stderr)
        return 2

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet1: 需求与目标
    ws = wb.create_sheet('需求与目标')
    ws.append(('参数', '值', '单位'))
    req = design.get('requirements') or {}
    for k, v in req.items():
        unit = v[1] if isinstance(v, (tuple, list)) and len(v) > 1 else ''
        val = v[0] if isinstance(v, (tuple, list)) else v
        ws.append((k, val, unit))

    # Sheet2: Cell 信息（计算结果）
    ws2 = wb.create_sheet('Cell信息')
    ws2.append(('参数', '值', '单位'))
    cell = [
        ('目标容量', design.get('design_capacity_Ah'), 'Ah'),
        ('设计容量', design.get('design_capacity_Ah'), 'Ah'),
        ('标称电压', design.get('nominal_voltage_V'), 'V'),
        ('能量', design.get('energy_Wh'), 'Wh'),
        ('重量', design.get('weight_g'), 'g'),
        ('重量能量密度', design.get('energy_density_Wh_kg'), 'Wh/kg'),
        ('体积能量密度', design.get('volumetric_density_Wh_L'), 'Wh/L'),
        ('循环寿命', design.get('cycle_life'), '次'),
        ('参数来源', design.get('source'), ''),
    ]
    for r in cell:
        ws2.append(r)

    # Sheet3: 完整参数集（41 项，11 节）
    ws3 = wb.create_sheet('完整参数集')
    ws3.append(('节', '参数', '值', '来源'))
    for section in (design.get('params41') or {}).get('sections') or []:
        for item in section['items']:
            ws3.append((section['name'], item['key'], item['value'], item['source']))

    # Sheet4: 工程参数明细（分项重量 + 目标校验）
    ws4 = wb.create_sheet('工程参数明细')
    ws4.append(('分项', '值', '单位'))
    for k, v in (design.get('detail') or {}).items():
        ws4.append((k, v, 'g'))
    target = design.get('target_check') or {}
    ws4.append(('目标密度', target.get('target_density_Wh_kg'), 'Wh/kg'))
    ws4.append(('密度达标', target.get('density_met'), ''))
    ws4.append(('重量上限', target.get('weight_limit_g'), 'g'))
    ws4.append(('重量达标', target.get('weight_met'), ''))

    for wsx in (ws, ws2, ws3, ws4):
        wsx.column_dimensions['A'].width = 24
        wsx.column_dimensions['B'].width = 30
        wsx.column_dimensions['C'].width = 20
        wsx.column_dimensions['D'].width = 12

    wb.save(out)
    print(json.dumps({"ok": True, "file": out}))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
